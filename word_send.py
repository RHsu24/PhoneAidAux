
import requests
from pathlib import Path
import json
import xlsxwriter
import pandas as pd
import os
import glob
from openpyxl import load_workbook, Workbook

def simple_split(namelist):
    #cuts out the number at the end
    word_extr = []
    rubbish = namelist.rpartition("_")
    word_extr = rubbish[0]

    return word_extr

def simple_truncate(namelist):
    new_dir = []
    rubbish = namelist.split("_",1)
    new_dir = rubbish[0]
    # print(new_name)
    return new_dir

def send_speech_analysis_request(
    speech_file_path,
    target_word,
    lang,
    age_cat,
    server_url="http://127.0.0.1:8000/analyze/",
):
    
    with open(speech_file_path, "rb") as f:
        files = {"speech_signal": f}
        data = {
            "target_word": target_word,
            "lang": lang,
            "age_cat": age_cat,
        }
        response = requests.post(server_url, files=files, data=data, timeout=120)
    if response.status_code == 200:
        return response.json()
    raise Exception(f"Error: {response.status_code}, {response.text}")


def main():
    #change rw_folder to path with segmented .wav files you want to analyse
    rw_folder = r'C:\Users\ricky\Desktop\Jobs\UCanSay66\First Week\Task1\Task1\PhonAid_to_be_scored\PhonAid_to_be_scored\663'
    os.chdir(rw_folder)

    #Excel spreadsheet initialisation, do only once
    workbook = xlsxwriter.Workbook('results.xlsx')
    worksheet = workbook.add_worksheet()
    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})
    worksheet.write('A1', 'Word', bold)
    worksheet.write('B1', 'word_id', bold)
    worksheet.write('C1', 'annotation', bold)
    worksheet.write('D1','0', bold)
    worksheet.write('E1','1', bold)
    worksheet.write('F1','2', bold)
    worksheet.write('G1','3', bold)
    worksheet.write('H1','4', bold)
    worksheet.write('I1','5', bold)
    worksheet.write('J1','6', bold)
    worksheet.write('K1','7', bold)
    worksheet.write('L1','8', bold)
    worksheet.write('M1','9', bold)
    worksheet.write('N1','10', bold)
    worksheet.write('O1','11', bold)
    worksheet.write('P1','12', bold)
    worksheet.write('Q1','13', bold)
    worksheet.write('R1','14', bold)
    worksheet.write('S1','15', bold)
    worksheet.write('T1','16', bold)
    worksheet.write('U1','17', bold)
    worksheet.write('V1','18', bold)
    worksheet.write('W1','19', bold)

    # Start from the first cell below the headers.
    row = 1
    col = 0
    word_count = 0
    for files in glob.glob("*.wav"):

        speech_file_path = files
        result = send_speech_analysis_request(
        speech_file_path=speech_file_path,
        target_word=simple_split(speech_file_path),
        lang="en_aus",
        age_cat="c",
        server_url="https://phoneaid-service-948892252068.australia-southeast1.run.app/analyze/",
        )

        test = json.dumps(result, ensure_ascii= False)
        parsing = json.loads(json.loads(test))
        phoneme = parsing["Phoneme"]
        exp = phoneme["Assess"]["Expected_aligned"]
        rec = phoneme["Assess"]["Recognized_aligned"]
        err = phoneme["Assess"]["Errors"]


        target_word = simple_split(speech_file_path)
        worksheet.write(row,col, target_word, bold)
        worksheet.write(row,col+1, speech_file_path, bold)
        worksheet.write(row,col+2, 'Expected', bold)
        worksheet.write(row+1,col+2, 'Recognised', bold)
        worksheet.write(row+2,col+2, 'Type', bold)
        for i, (e, r, tag) in enumerate(zip(exp, rec, err), start=1):
            #print(f"{i:02d}. {e} -> {r} [{tag}]")
            worksheet.write(row,col+3, e)
            worksheet.write(row+1,col+3, r)
            worksheet.write(row+2,col+3,tag)  
            col+=1
        col=0
        row+=3
        word_count += 1

    workbook.close()
    print(f"Speech Analysis Done! There were {word_count} words analysed")

    # Format XLSX sheet to look nicer

    print("Beginning Excel Sheet Cleanup")
    # Merge cells to make output file look nicer
    row = 2
    col = 1
    
    wb = load_workbook(filename = 'results.xlsx')
    ws = wb.active
    for row in range(2, 3*word_count+1,3):
        ws.merge_cells(start_row=row,end_row=row+2,start_column=col,end_column=col)
        ws.merge_cells(start_row=row,end_row=row+2,start_column=col+1,end_column=col+1)  

    wb.save('results.xlsx')
    print("Excel Sheet Prettified!")
if __name__ == '__main__':

    welcome_msg = "This is here to show it is running\n"
    print(welcome_msg)
    
    main()